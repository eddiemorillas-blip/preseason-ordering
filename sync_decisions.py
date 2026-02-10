"""
Sync SHIP/CANCEL decisions from Scarpa and La Sportiva spreadsheets
to the preseason ordering database.

Usage:
  python3 sync_decisions.py

Requires DATABASE_URL in .env.local (or .env) file.
"""
import psycopg2
import openpyxl
import os
import sys

# Load environment from .env.local or .env
try:
    from dotenv import load_dotenv
    env_path = os.path.join(os.path.dirname(__file__), '.env.local')
    if os.path.exists(env_path):
        load_dotenv(env_path)
    else:
        load_dotenv()  # Try .env
except ImportError:
    pass  # dotenv not installed, rely on environment

DB_URL = os.environ.get('DATABASE_URL')
if not DB_URL:
    print("ERROR: DATABASE_URL not found")
    print("Add DATABASE_URL to .env.local or install python-dotenv: pip install python-dotenv")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def connect():
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    return conn

def get_brands(cur):
    """Show all brands in the database"""
    cur.execute("SELECT id, name FROM brands ORDER BY id")
    brands = cur.fetchall()
    print("=== BRANDS IN DATABASE ===")
    for b in brands:
        print(f"  ID {b[0]}: {b[1]}")
    return {name.lower(): bid for bid, name in brands}

def get_locations(cur):
    """Show all locations"""
    cur.execute("SELECT id, name FROM locations ORDER BY id")
    locs = cur.fetchall()
    print("\n=== LOCATIONS ===")
    for l in locs:
        print(f"  ID {l[0]}: {l[1]}")
    return {name.lower(): lid for lid, name in locs}

def find_orders_with_items(cur, brand_id, location_id, ship_date_prefix):
    """Find orders matching brand/location/ship date and return their items with UPCs"""
    cur.execute("""
        SELECT o.id, o.order_number, o.ship_date,
               oi.id AS item_id, p.upc, p.name AS product_name,
               p.size, oi.quantity, oi.adjusted_quantity, oi.vendor_decision
        FROM orders o
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_id = p.id
        WHERE o.brand_id = %s
          AND o.location_id = %s
          AND o.order_number LIKE %s
        ORDER BY o.order_number, p.name, p.size
    """, (brand_id, location_id, f"{ship_date_prefix}%"))
    return cur.fetchall()

def sync_scarpa(cur, brand_map, loc_map):
    """Sync Scarpa decisions from spreadsheet"""
    print("\n" + "="*60)
    print("SYNCING SCARPA DECISIONS")
    print("="*60)

    filepath = os.path.join(BASE_DIR, "Scarpa_OrderReview.xlsx")
    if not os.path.exists(filepath):
        print(f"ERROR: {filepath} not found")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Order Review"]

    # Build decisions from spreadsheet: {(upc, location): decision}
    decisions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:  # skip empty rows
            break
        # Actual columns: SO#(0), Location(1), ItemCode(2), Description(3), Size(4),
        #                  Qty(5), UnitPrice(6), UPC(7), OnHand(8), Decision(9), Notes(10)
        so = row[0]
        location = str(row[1]).strip() if row[1] else None
        product = row[3]
        size = row[4]
        qty = row[5]
        upc = str(row[7]).strip() if row[7] else None
        decision = str(row[9]).strip() if row[9] else None

        if upc and decision and location:
            decisions.append({
                'upc': upc,
                'decision': decision,
                'qty': qty,
                'location': location,
                'product': product,
                'size': size
            })

    print(f"  Read {len(decisions)} items from spreadsheet")

    # Map location names from spreadsheet to DB location IDs
    loc_mapping = {
        'ogden': 'ogden',
        'slc': 'salt lake',
        'south main': 'south main',
    }

    # Get Scarpa brand ID
    scarpa_id = None
    for name, bid in brand_map.items():
        if 'scarpa' in name:
            scarpa_id = bid
            break

    if not scarpa_id:
        print("ERROR: Scarpa brand not found in database!")
        return

    print(f"  Scarpa brand_id = {scarpa_id}")

    # For each location, find the FEB26 orders and match UPCs
    updated = 0
    not_found = 0
    already_set = 0

    for loc_label, db_loc_key in loc_mapping.items():
        loc_id = loc_map.get(db_loc_key)
        if not loc_id:
            print(f"  WARNING: Location '{db_loc_key}' not found in DB")
            continue

        # Get all order items for Scarpa FEB26 at this location
        items = find_orders_with_items(cur, scarpa_id, loc_id, "FEB26-SCA")

        # Build a UPC->item_id map for this location
        upc_to_item = {}
        for row in items:
            order_id, order_num, ship_date, item_id, upc, pname, size, orig_qty, adj_qty, vd = row
            if upc:
                upc_to_item[upc.strip()] = {
                    'order_id': order_id,
                    'item_id': item_id,
                    'product': pname,
                    'size': size,
                    'current_decision': vd
                }

        print(f"\n  Location: {loc_label.upper()} (id={loc_id})")
        print(f"    DB items found: {len(upc_to_item)}")

        # Match spreadsheet decisions to DB items
        loc_decisions = [d for d in decisions if d['location'].lower() == loc_label.lower()]
        print(f"    Spreadsheet items: {len(loc_decisions)}")

        for d in loc_decisions:
            upc = d['upc']
            item = upc_to_item.get(upc)
            if not item:
                print(f"    NOT FOUND in DB: {upc} ({d['product']} {d['size']})")
                not_found += 1
                continue

            if item['current_decision']:
                print(f"    ALREADY SET: {upc} ({item['product']} {item['size']}) = {item['current_decision']}")
                already_set += 1
                continue

            # Determine vendor_decision and receipt_status
            if d['decision'].upper() == 'CANCEL':
                vendor_decision = 'cancel'
                adj_qty = 0
                receipt_status = 'cancelled'
            else:  # SHIP
                vendor_decision = 'ship'
                adj_qty = d['qty'] if d['qty'] else 1
                receipt_status = 'pending'

            cur.execute("""
                UPDATE order_items
                SET vendor_decision = %s,
                    adjusted_quantity = %s,
                    receipt_status = %s
                WHERE id = %s
            """, (vendor_decision, adj_qty, receipt_status, item['item_id']))

            updated += 1
            print(f"    OK: {upc} ({item['product']} {item['size']}) → {vendor_decision.upper()}")

    print(f"\n  SCARPA SUMMARY:")
    print(f"    Updated: {updated}")
    print(f"    Not found in DB: {not_found}")
    print(f"    Already set: {already_set}")
    return updated


def sync_lasportiva(cur, brand_map, loc_map):
    """Sync La Sportiva decisions from spreadsheet"""
    print("\n" + "="*60)
    print("SYNCING LA SPORTIVA DECISIONS")
    print("="*60)

    filepath = os.path.join(BASE_DIR, "LaSportiva_REVISED.xlsx")
    if not os.path.exists(filepath):
        print(f"ERROR: {filepath} not found")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the right sheet - it might be "Order Review" or "REVISE HERE"
    sheet_names = wb.sheetnames
    print(f"  Available sheets: {sheet_names}")

    ws = None
    for name in ['REVISE HERE', 'Order Review']:
        if name in sheet_names:
            ws = wb[name]
            print(f"  Using sheet: {name}")
            break

    if ws is None:
        ws = wb[sheet_names[0]]
        print(f"  Using first sheet: {sheet_names[0]}")

    # La Sportiva format: first few rows are headers
    # Find the header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if row and any(str(c).lower().strip() in ('upc', 'barcode', 'item upc') for c in row if c):
            header_row = row_idx
            break

    if not header_row:
        # Try a fixed row based on what we know
        header_row = 7
        print(f"  Using assumed header row: {header_row}")
    else:
        print(f"  Found header at row: {header_row}")

    # Read headers
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value).strip().lower() if cell.value else '')

    print(f"  Headers: {headers}")

    # Find column indices
    upc_col = None
    decision_col = None
    adj_qty_col = None
    location_col = None
    product_col = None
    size_col = None

    for i, h in enumerate(headers):
        if 'upc' in h or 'barcode' in h:
            upc_col = i
        elif 'ship' in h and 'cancel' in h:
            # "please mark items as ship / cancel below"
            decision_col = i
        elif 'decision' in h or 'status' in h:
            decision_col = i
        elif 'quantiy adjustments' in h or 'quantity adjustment' in h or ('adjusted' in h and 'quant' in h):
            adj_qty_col = i
        elif ('location' in h or 'facility' in h or 'store' in h) and 'ship' in h:
            # "ship to location"
            location_col = i
        elif 'location' in h or 'facility' in h or 'store' in h:
            location_col = i
        elif 'item name' in h:
            product_col = i
        elif ('product' in h or 'name' in h or 'description' in h) and product_col is None:
            product_col = i
        elif 'size' in h:
            size_col = i

    # Fallback: hardcode known column positions if detection missed them
    if upc_col is None:
        upc_col = 14  # "upc code"
    if decision_col is None:
        decision_col = 20  # "please mark items as ship / cancel below"
    if adj_qty_col is None:
        adj_qty_col = 19  # "quantiy adjustments (enter new quanity)"
    if location_col is None:
        location_col = 2  # "ship to location"
    if product_col is None:
        product_col = 12  # "item name"

    print(f"  Column mapping: upc={upc_col}, decision={decision_col}, adj_qty={adj_qty_col}, location={location_col}, product={product_col}")

    # Get La Sportiva brand ID
    lsp_id = None
    for name, bid in brand_map.items():
        if 'sportiva' in name:
            lsp_id = bid
            break

    if not lsp_id:
        print("ERROR: La Sportiva brand not found in database!")
        return

    print(f"  La Sportiva brand_id = {lsp_id}")

    # Location mapping for La Sportiva spreadsheet
    loc_mapping = {
        'ogden': 'ogden',
        'salt lake city': 'salt lake',
        'salt lake': 'salt lake',
        'slc': 'salt lake',
        'millcreek': 'south main',
        'south main': 'south main',
        'soma': 'south main',
        'front climbing club': None,  # skip - not a specific location
    }

    # Read all data rows
    decisions = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        if not row or not row[upc_col]:
            continue

        upc = str(row[upc_col]).strip() if row[upc_col] else None
        decision = str(row[decision_col]).strip() if row[decision_col] else None
        adj_qty = row[adj_qty_col] if adj_qty_col is not None and row[adj_qty_col] is not None else None
        location = str(row[location_col]).strip() if location_col is not None and row[location_col] else None
        product = str(row[product_col]).strip() if product_col is not None and row[product_col] else ''
        size = str(row[size_col]).strip() if size_col is not None and row[size_col] else ''

        if upc and decision and location:
            # Clean up UPC - remove .0 if it's a float
            if upc.endswith('.0'):
                upc = upc[:-2]

            decisions.append({
                'upc': upc,
                'decision': decision,
                'adj_qty': adj_qty,
                'location': location,
                'product': product,
                'size': size
            })

    print(f"  Read {len(decisions)} items from spreadsheet")

    # Get ALL La Sportiva Spring 2026 orders with items
    # The pending report covers JAN26 through JUN26 ship dates
    cur.execute("""
        SELECT o.id, o.order_number, o.ship_date, o.location_id,
               oi.id AS item_id, p.upc, p.name AS product_name,
               p.size, oi.quantity, oi.adjusted_quantity, oi.vendor_decision
        FROM orders o
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_id = p.id
        WHERE o.brand_id = %s
          AND o.order_number LIKE 'FEB26-LAS%%'
        ORDER BY o.location_id, o.order_number, p.name, p.size
    """, (lsp_id,))

    all_items = cur.fetchall()
    print(f"  FEB26-LAS items in DB: {len(all_items)}")

    # Also check JAN26, MAR26 etc. since the pending report spans multiple months
    for prefix in ['JAN26-LAS', 'MAR26-LAS', 'APR26-LAS', 'MAY26-LAS', 'JUN26-LAS']:
        cur.execute("""
            SELECT o.id, o.order_number, o.ship_date, o.location_id,
                   oi.id AS item_id, p.upc, p.name AS product_name,
                   p.size, oi.quantity, oi.adjusted_quantity, oi.vendor_decision
            FROM orders o
            JOIN order_items oi ON o.id = oi.order_id
            JOIN products p ON oi.product_id = p.id
            WHERE o.brand_id = %s
              AND o.order_number LIKE %s
            ORDER BY o.location_id, o.order_number, p.name, p.size
        """, (lsp_id, f"{prefix}%"))
        extra = cur.fetchall()
        if extra:
            print(f"  {prefix} items in DB: {len(extra)}")
            all_items.extend(extra)

    # Build a (upc, location_id) → list of items map
    # Multiple orders may have the same UPC at the same location (different ship dates)
    upc_loc_items = {}
    for row in all_items:
        order_id, order_num, ship_date, loc_id, item_id, upc, pname, size, orig_qty, adj_qty, vd = row
        if upc:
            key = (upc.strip(), loc_id)
            if key not in upc_loc_items:
                upc_loc_items[key] = []
            upc_loc_items[key].append({
                'order_id': order_id,
                'order_number': order_num,
                'item_id': item_id,
                'product': pname,
                'size': size,
                'orig_qty': orig_qty,
                'current_decision': vd
            })

    print(f"  Unique (UPC, location) combos in DB: {len(upc_loc_items)}")

    updated = 0
    not_found = 0
    already_set = 0
    multi_match = 0

    for d in decisions:
        # Map spreadsheet location to DB location ID
        loc_key = d['location'].lower().strip()
        if loc_key not in loc_mapping:
            print(f"    UNKNOWN LOCATION: '{d['location']}' for {d['upc']}")
            not_found += 1
            continue

        db_loc_name = loc_mapping[loc_key]
        if db_loc_name is None:
            # "Front Climbing Club" or similar - skip
            not_found += 1
            continue

        loc_id = loc_map.get(db_loc_name)
        if not loc_id:
            print(f"    LOCATION NOT IN DB: {db_loc_name}")
            not_found += 1
            continue

        upc = d['upc']
        items = upc_loc_items.get((upc, loc_id), [])

        if not items:
            # Try without leading zeros or with different formatting
            not_found += 1
            if not_found <= 10:
                print(f"    NOT FOUND: {upc} @ {d['location']} ({d['product']} {d['size']})")
            continue

        # Determine vendor_decision
        decision_text = d['decision'].lower()
        if 'cancel' in decision_text:
            vendor_decision = 'cancel'
            adj_qty = 0
            receipt_status = 'cancelled'
        elif 'keep' in decision_text or 'b/o' in decision_text or 'back' in decision_text:
            vendor_decision = 'keep_open_bo'
            adj_qty = d['adj_qty'] if d['adj_qty'] is not None else None
            receipt_status = 'backordered'
        elif 'ship' in decision_text:
            vendor_decision = 'ship'
            adj_qty = d['adj_qty'] if d['adj_qty'] is not None else None
            receipt_status = 'pending'
        else:
            print(f"    UNKNOWN DECISION: '{d['decision']}' for {upc}")
            continue

        # Update ALL matching items (same UPC can be in multiple orders for same location)
        for item in items:
            if item['current_decision']:
                already_set += 1
                continue

            final_qty = adj_qty if adj_qty is not None else item['orig_qty']

            cur.execute("""
                UPDATE order_items
                SET vendor_decision = %s,
                    adjusted_quantity = %s,
                    receipt_status = %s
                WHERE id = %s
            """, (vendor_decision, final_qty, receipt_status, item['item_id']))

            updated += 1
            if len(items) > 1:
                multi_match += 1

    if not_found > 10:
        print(f"    ... and {not_found - 10} more not found")

    print(f"\n  LA SPORTIVA SUMMARY:")
    print(f"    Updated: {updated}")
    print(f"    Not found in DB: {not_found}")
    print(f"    Already set: {already_set}")
    print(f"    Multi-order matches: {multi_match}")
    return updated


def main():
    conn = connect()
    cur = conn.cursor()

    try:
        # Step 1: Show brands and locations
        brand_map = get_brands(cur)
        loc_map = get_locations(cur)

        # Step 2: Sync Scarpa
        scarpa_count = sync_scarpa(cur, brand_map, loc_map)

        # Step 3: Sync La Sportiva
        lsp_count = sync_lasportiva(cur, brand_map, loc_map)

        # Step 4: Commit
        total = (scarpa_count or 0) + (lsp_count or 0)
        if total > 0:
            print(f"\n{'='*60}")
            print(f"COMMITTING {total} total updates...")
            conn.commit()
            print("DONE! All changes committed.")
        else:
            print("\nNo changes to commit.")
            conn.rollback()

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    main()
